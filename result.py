from openpyxl import load_workbook
wb=load_workbook('EmployeeData.xlsx')
ws=wb.active
male_coun=0
for row in range(2,ws.max_row+1):
    gender=ws.cell(row=row,column=4).value 
    if gender=='Male':
        male_coun+=1 
print('1.uzd')
print(male_coun)



ages_accounting=[]
for row in range(2,ws.max_row + 1):
    department=ws.cell(row=row,column=3).value
    age=ws.cell(row=row,column=5).value 
    if department=='Accounting' and isinstance(age,(int, float)):
        ages_accounting.append(age)
max_age_accounting=max(ages_accounting) 
print('2.uxd')
print(max_age_accounting)



total_age_it = 0
count_it_employees = 0
for row in range(2, ws.max_row + 1):
    department = ws.cell(row=row, column=3).value  
    age = ws.cell(row=row, column=5).value       
    if department == 'Finance' and isinstance(age, (int, float)):
        total_age_it += age
        count_it_employees += 1 
if count_it_employees > 0:
    average_age_it = total_age_it / count_it_employees  
    average_age_int = round(average_age_it)  
print('3.uzd')
print(average_age_int)



count = 0
for row in range(2, ws.max_row + 1):
    salary = ws.cell(row=row, column=6).value
    if 100000 < salary < 250000:
        count += 1
print('4.uzd')
print(count)
wb.close()


max_salary = 0
max_salary_row =0
department_with_max_salary = ''
for row in range(2, ws.max_row + 1):
    salary = ws.cell(row=row, column=6).value
    department = ws.cell(row=row, column=3).value
    if salary > max_salary:
        max_salary = salary
        max_salary_row = row
        department_with_max_salary = department
print('5.uzd')
print(department_with_max_salary)
wb.close()


import csv
with open('data.csv','r') as file:
    csv_reader = csv.DictReader(file)
    mining_values=[]
    for row in csv_reader:
        if row['industry']=='Mining':
            mining_values.append(float(row['value']))
    if mining_values:
        max_value=max(mining_values)
        print('6.uzd') 
        print(max_value)


import csv
with open('data.csv','r') as file:
    csv_reader = csv.DictReader(file)
    filtered_values=[]
    for row in csv_reader:
        if row['line_code']=='C0300.02':
            filtered_values.append(row['value'])
print('7.uzd')
print(len(filtered_values))


import csv
with open('data.csv', mode='r') as file:
    csv_reader = csv.DictReader(file)
    agriculture_values = []
    for row in csv_reader:
        if row['industry'] == 'Agriculture':
            agriculture_values.append(float(row['value']))
    if agriculture_values:
        average_value = sum(agriculture_values) / len(agriculture_values)
        rounded_average = round(average_value)
print('8.uzd')
print(rounded_average)



import csv
with open('data.csv', mode='r') as file:
    csv_reader = csv.DictReader(file)
    construction_values = []
    for row in csv_reader:
        if row['industry'] == 'Construction':
            construction_values.append(float(row['value']))
    
    # Atlasām TOP 3 vērtības, sakārtojot sarakstu dilstošā secībā
    top_3_values = sorted(construction_values, reverse=True)[:3]
    
    # Nosakām minimālo vērtību no TOP 3 vērtībām
    if top_3_values:
        min_top_3_value = min(top_3_values)
        print('9.uzd')
        print(min_top_3_value)














import csv
with open('data.csv', mode='r') as file:
    csv_reader = csv.DictReader(file)
    insurance_values = []
    for row in csv_reader:
        if row['industry'] == 'Insurance':
            insurance_values.append(float(row['value']))
    every_second_value = insurance_values[1::2]
    total_sum = sum(every_second_value)
print('10.uzd')
print(total_sum)








